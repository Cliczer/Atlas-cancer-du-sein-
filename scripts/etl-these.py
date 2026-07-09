# Extraction d'une base d'études (base_etudes.json) depuis le tableau Excel de
# thèse « cancer du sein ». Feuille « Données », en-têtes sur 2 lignes.
#   Usage : python3 etl-these.py <these.xlsx> [sortie.json]
# Étiquette les bras par différentiel d'intervention (chirurgie/RT/chimio…),
# déduit temps + famille des libellés d'issues, et nettoie titres/niveaux.
import openpyxl, json, re, sys
F   = sys.argv[1] if len(sys.argv) > 1 else "these.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "base_new.json"
wb=openpyxl.load_workbook(F, data_only=True)
ws=wb["Données"]
rows=list(ws.iter_rows(values_only=True))
h1=[("" if c is None else str(c).strip()) for c in rows[0]]
h2=[("" if c is None else str(c).strip()) for c in rows[1]]
def cn(i):
    a=h1[i] if i<len(h1) else ""; b=h2[i] if i<len(h2) else ""
    return (a+"/"+b).strip("/") if(a and b) else (a or b) or f"col{i}"
COLS={cn(i):i for i in range(ws.max_column)}
def idx(name): return COLS.get(name)

# index des colonnes clés
C_REF,C_OBJ=0,1
C_TYPE,C_ECH,C_SUIVI=3,7,8
C_NP=12
# Uniquement les critères d'ÉLIGIBILITÉ, clés = clés du dictionnaire vocabulaire.json.
# (Les interventions servent à détecter les bras — pas besoin comme critères.)
CRIT={ "T":14,"N":15,"M":16,"RE":17,"RP":18,"HER2":19,"Ki67":20,
       "Marges":21,"Mutation":24 }
C_RT=27; C_N=15

# colonnes d'issues (résultats), 34..92 ; on exclut les stats RR/IC/HR/RRA
EXCLUDE_NAMES={"RR","IC à 95%","IC 95%","RRA"}
# Expansion des acronymes cliniques standards (libellés lisibles).
ACRO={ "os":"Survie globale","dfs":"Survie sans maladie","bcss":"Survie spécifique au cancer du sein",
       "bcm":"Mortalité par cancer du sein","bcd":"Décès par cancer du sein",
       "lrr":"Récidive loco-régionale","lr":"Récidive locale","cbc":"Cancer du sein controlatéral",
       "ibr":"Récidive homolatérale","distant-dfs":"Survie sans métastase",
       "death from all causes":"Décès toutes causes","crc":"Réponse complète clinique",
       "crp":"Réponse complète pathologique" }
def expanse(fam):
    return ACRO.get(fam.strip().lower(), fam)
def cleanname(nm):
    return re.sub(r'^Issues\s*/\s*','',nm).strip()
def is_outcome(i):
    if i<34: return False
    nm=cleanname(cn(i))
    if nm in EXCLUDE_NAMES: return False
    if re.match(r'^col\d+$', nm): return False   # colonnes sans en-tête
    toks=re.split(r'[\s/]+', nm)
    if any(t in ("RR","HR","IC") for t in toks): return False   # colonnes de statistiques
    return True
OUTCOLS=[i for i in range(34, ws.max_column) if is_outcome(i)]

def val(r,i):
    if i is None or i>=len(r): return None
    v=r[i]
    return None if v in (None,"") else v
def neutre(v):
    return v is None or str(v).strip().lower() in ("nc","-1","","n/a","nan","nr")

def parse_pct(v):
    s=str(v)
    m=re.search(r'-?\d+(?:[.,]\d+)?', s)
    if not m: return None
    n=float(m.group(0).replace(',','.'))
    pctsign='%' in s
    if not pctsign and 0<abs(n)<=1: n=n*100
    return round(n,1)

def fam_temps(nm):
    # "OS à 10A" -> ("OS",10) ; "BCM à 20A" -> ("BCM",20) ; "Récidive à 10A" -> ("Récidive",10)
    t=None
    m=re.search(r'à\s*(\d+)\s*A\b', nm) or re.search(r'(\d+)\s*ans?', nm, re.I) or re.search(r'à\s*(\d+)\s*A', nm)
    if m: t=int(m.group(1))
    fam=re.sub(r'\s*à?\s*\d+\s*A\b','',nm).replace('  ',' ').strip()
    fam=re.sub(r'\s*à\s*\d+\s*ans?','',fam, flags=re.I).strip()
    return (fam or nm, t)
def mesure_nom(nm):
    return re.sub(r'à\s*(\d+)\s*A\b', r'à \1 ans', nm)

# grouper en études
data=rows[2:]; studies=[]; cur=None
for r in data:
    if not r: continue
    if r[0] not in (None,""):
        cur={"rows":[r]}; studies.append(cur)
    elif cur is not None and any(c not in (None,"") for c in r):
        cur["rows"].append(r)

def clean_np(v):
    # Le niveau de preuve doit être un jeton court (1, 2, 3, 4, III…). Certaines
    # lignes contiennent un paragraphe (décalage de colonnes) : on l'écarte.
    s=str(v or "").strip()
    if not s: return ""
    if len(s)>6: return ""
    if re.fullmatch(r'[0-9IViv]{1,4}[abAB]?', s): return s.upper() if any(c in s for c in "iIvV") else s
    return ""

def fallback_titre(ref, i):
    # Objectif absent → titre « Auteur et al. (année) » déduit de la référence.
    s=re.sub(r'^\s*\(\d+\)\s*', '', str(ref or "").strip())          # retire "(9) "
    if not s: return f"Étude {i+1}"
    # nom de famille (avec particules « van der », « de la »…) suivi des initiales
    m=re.match(r'((?:(?i:van|von|de|der|den|du|le|la|di|da|dos|del|el)\s+)*[A-ZÀ-Ÿ][a-zà-ÿ]+(?:-[A-ZÀ-Ÿ][a-zà-ÿ]+)?)\s+[A-ZÀ-Ÿ]{1,3}[,.\s]', s)
    nom=(m.group(1).strip() if m else re.split(r'[,\s]', s)[0]).strip(" ,.")
    nom=nom[:1].upper()+nom[1:] if nom else ""
    my=re.search(r'\b(19|20)\d{2}\b', str(ref or ""))
    an=my.group(0) if my else ""
    if not nom: return f"Étude {i+1}"
    return f"{nom} et al." + (f" ({an})" if an else "")

# --- Étiquetage des bras par différentiel d'intervention -------------------
# Chaque étude compare des bras qui diffèrent sur UNE ou plusieurs interventions
# (chirurgie, RT, chimio…). Les lignes suivantes ne re-remplissent que les
# colonnes qui changent → on propage (forward-fill) le profil de la 1re ligne,
# puis on étiquette chaque bras par le plus petit ensemble d'axes qui le rend
# unique. Bien plus fidèle que « Avec/Sans RT » systématique.
INTV_NAME={25:"ChirM",26:"ChirAx",27:"RT",28:"CNA",29:"CTratt",30:"CTadj",31:"Immuno",32:"Hormo"}
INTV_PRIORITY=[26,25,27,30,32,28,31,29]  # ChirAx, ChirM, RT, CTadj, Hormo, CNA, Immuno, CTratt

def norm(x): return (str(x or "")).strip().lower().rstrip(" ,.").strip()

def short_val(ci, v):
    x=norm(v)
    if not x or x in ("nc","n/a","nan"): return ""
    if ci==27:  # RT
        if x.startswith(("non","aucun","pas","ort","ørt")) or x=="0": return "sans RT"
        if "hypofraction" in x: return "RT hypofractionnée"
        if "standard" in x or "normofraction" in x or "conventionn" in x: return "RT standard"
        if any(t in x for t in ("cmi","supra","sus-clav","clavicul")): return "RT + aires ggl"
        return "avec RT"
    if ci==26:  # Chirurgie axillaire
        if x.startswith(("non","aucun","abst","pas")): return "sans curage"
        toks=re.split(r'[ ,]+', x)
        hasGS = "gs" in toks or "sentinel" in x or "sentinelle" in x
        hasCA = "ca" in toks or "axillaire" in x or "curage" in x
        if hasGS and hasCA: return "GS + curage"
        if hasGS: return "GS seul"
        if hasCA: return "curage axillaire"
        return v.strip()
    if ci==25:  # Chirurgie mammaire
        if "radical" in x: return "mastectomie radicale"
        if "tssm" in x or "épargn" in x or "epargn" in x or "reconstru" in x: return "mastectomie + reconstruction"
        if x.startswith("mt") or "totale" in x: return "mastectomie totale"
        if x.startswith("mp") or "partiel" in x or "conservat" in x or "tumorect" in x: return "chir. conservatrice"
        return v.strip()
    if ci==30: return "avec chimio" if x.startswith("oui") else ("sans chimio" if x.startswith("non") else "")
    if ci==28: return "avec chimio néoadj." if x.startswith("oui") else ("sans chimio néoadj." if x.startswith("non") else "")
    if ci==31: return "avec immunothérapie" if x.startswith("oui") else ("sans immunothérapie" if x.startswith("non") else "")
    if ci==32: return "sans hormonothérapie" if x.startswith("non") else "avec hormonothérapie"
    return v.strip()

def combo(prof, axes):
    parts=[short_val(c, prof.get(c,"")) for c in axes]
    return " · ".join([p for p in parts if p])

def label_arms(profs):
    # profs : liste de profils (dict {ci:val}) forward-fillés, un par bras du groupe
    n=len(profs)
    if n==1:
        p=profs[0]
        parts=[short_val(c,p.get(c,"")) for c in INTV_PRIORITY if short_val(c,p.get(c,""))]
        return [" · ".join(parts[:2]) or "Résultat global"]
    chosen=[]
    for c in INTV_PRIORITY:
        vals={short_val(c,p.get(c,"")) for p in profs}
        if len(vals)>1:
            chosen.append(c)
            if len({combo(p,chosen) for p in profs})==n: break
    labels=[combo(p,chosen) for p in profs]
    seen={}; out=[]
    for i,l in enumerate(labels):
        l=l or ("Bras "+str(i+1))
        if l in seen: seen[l]+=1; l=l+" ("+str(seen[l])+")"
        else: seen[l]=1
        out.append(l)
    return out

def subg_of(v):
    # sous-groupe = valeur N courte (pN0, pN1-3, N4+…). Rejette les longues listes
    # d'inclusion (« N0, N1, N2, pN0,… ») qui ne sont pas un sous-groupe.
    x=str(v or "").strip()
    if not x or neutre(x): return None
    if len(x)>10 or x.count(",")>=2: return None
    return x

etudes=[]
for _i, s in enumerate(studies):
    rws=s["rows"]; r0=rws[0]
    # métadonnées
    ref=str(val(r0,C_REF) or "").strip()
    obj=str(val(r0,C_OBJ) or "").strip()
    typ=str(val(r0,C_TYPE) or "").strip()
    np_=val(r0,C_NP); np_=("" if np_ is None else str(np_).replace(".0","").strip())
    # criteres bruts (première ligne ; NC -> absent)
    criteres={}
    for k,ci in CRIT.items():
        v=val(r0,ci)
        if v is not None and not neutre(v): criteres[k]=str(v).strip()
    # profils d'intervention forward-fillés (un par ligne)
    profils=[]; last={}
    for r in rws:
        cur=dict(last)
        for ci in INTV_NAME:
            v=val(r,ci)
            if v is not None and norm(v) and norm(v)!="nc": cur[ci]=str(v).strip()
        last=cur; profils.append(cur)
    # comparaisons : par colonne d'issue ; bras = lignes ayant une valeur ; sous-groupe via N courant
    comps=[]
    for ci in OUTCOLS:
        nm=cleanname(cn(ci))
        entries=[]; subg=None       # (index_ligne, sousgroupe, valeur)
        for ri,r in enumerate(rws):
            sg=subg_of(val(r,C_N))
            if sg is not None: subg=sg
            v=val(r,ci)
            if v is None: continue
            p=parse_pct(v)
            if p is None: continue
            entries.append((ri, subg, p))
        if not entries: continue
        rawfam,temps=fam_temps(nm)
        fam=expanse(rawfam)
        base_nom=fam+(f" à {temps} ans" if temps is not None else "")
        bysg={}
        for ri,sg,p in entries:
            bysg.setdefault(sg,[]).append((ri,p))
        multi=len(bysg)>1
        for sg,arms in bysg.items():
            # >4 bras = effondrement de sous-groupes non identifiés (ex. groupes de
            # risque Oncotype hors colonne N) : non interprétable → on écarte.
            if len(arms)>4: continue
            mnom=base_nom+((" — "+sg) if (multi and sg) else "")
            labels=label_arms([profils[ri] for ri,_ in arms])
            bras=[{"label": labels[k], "valeur": p} for k,(ri,p) in enumerate(arms)]
            comp={"mesure":mnom,"unite":"","bras":bras}
            if temps is not None: comp["temps"]=temps
            comp["famille"]=fam
            comps.append(comp)
    titre=(obj[:1].upper()+obj[1:]) if obj else fallback_titre(ref, _i)
    et={"titre": titre,
        "auteurs":"", "reference": ref, "objectif": obj,
        "niveau_preuve": clean_np(np_), "type_etude": typ,
        "lien":"", "criteres":criteres, "comparaisons":comps,
        "importance":0}
    etudes.append(et)

out={"etudes":etudes}
open(OUT,"w").write(json.dumps(out,ensure_ascii=False,indent=4)+"\n")
print("Études écrites:", len(etudes), "→", OUT)
tot_comp=sum(len(e["comparaisons"]) for e in etudes)
print("Comparaisons totales:", tot_comp)
# aperçu étude 0 et 1
for si in (0,1):
    e=etudes[si]; print(f"\n#{si} ref={e['reference'][:40]} | np={e['niveau_preuve']} | criteres={list(e['criteres'].keys())}")
    for c in e["comparaisons"][:5]:
        print("   ", c["mesure"], "| fam",c.get("famille"),"t",c.get("temps"),"|", [(b['label'],b['valeur']) for b in c['bras']])
